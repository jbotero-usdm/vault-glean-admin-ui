#!/usr/bin/env python3
"""
Apply proper GxP security model to ABC Biopharma demo vault.

SECURITY MODEL
==============

Three admin users get access to EVERYTHING (no filtering):
  - Jennell Botero   (jbotero@usdm.com)
  - Brittany Walker  (bwalker@usdm.com)         [look up actual email at runtime]
  - Hovsep Kirikian  (hkirikian@usdm.com)

Other users get role-appropriate access via lifecycle states:

DOCUMENTS:
  Draft / In Review     → Author/Owner only (created_by__v)
  Approved / Effective  → All active Vault users (Viewer role)
  Final / Closed        → All active Vault users (Viewer role)
  Superseded / Obsolete → All active Vault users (Viewer role)

QMS OBJECT RECORDS (deviation, CAPA, audit, change_control, etc.):
  Open / Draft / In Progress  → Owner + admins only
  Closed / Effective / Approved → All users via DAC

  Owner__sys is always set to the record creator if missing.

WHAT THIS SCRIPT DOES
=====================
1. Reads current state of Vault: users, groups, docs, objects
2. Identifies admins by email matching (case-insensitive)
3. For each document:
   - Always assigns the 3 admins to owner__v role (so they see everything)
   - For docs in non-draft states, assigns all active users to viewer__v role
4. For each object record:
   - Ensures owner__sys is set (defaults to created_by__v if missing)
   - For closed/approved/effective records, grants DAC access to all active users
     via the Viewer role
5. Writes a detailed Excel report showing what changed and why

USAGE
=====
  # First, dry-run to see what would change
  python vault_security_apply_gxp.py --dry-run

  # Then apply for real
  python vault_security_apply_gxp.py --apply

  # Audit only — no changes, just show current state
  python vault_security_apply_gxp.py --audit-only
"""

import argparse
import csv
import datetime
import os
import sys
import time
from collections import Counter, defaultdict

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

load_dotenv()

VAULT_DNS = os.getenv("VAULT_DNS", "partnersi-usdm-quality.veevavault.com")
VAULT_USER = os.getenv("VAULT_USERNAME")
VAULT_PASS = os.getenv("VAULT_PASSWORD")
VAULT_API = os.getenv("VAULT_API_VERSION", "v24.3")
REQUEST_TIMEOUT = int(os.getenv("REQUEST_TIMEOUT", "60"))

# ----- ADMIN USERS — these get access to everything -----
# Edit ONLY this list if admins change. Match is case-insensitive on email or
# (last_name + first_initial) as a fallback when emails don't match exactly.
ADMIN_HINTS = [
    {"email_contains": "jbotero", "last_name": "botero", "first_name": "jennell"},
    {"email_contains": "bwalker", "last_name": "walker", "first_name": "brittany"},
    {"email_contains": "hkirikian", "last_name": "kirikian", "first_name": "hovsep"},
]

# Object types covered (QMS + related)
OBJECT_TYPES = [
    "deviation__v", "change_control__v", "nonconformance__v", "audit__qdm",
    "investigation__qdm", "capa_action__qdm", "effectiveness_check__qdm",
    "finding__v", "complaint__v", "disposition_plan__v",
    "quality_batch__v", "quality_material__v",
    "facility__v", "product__v", "product_family__v",
]

# Lifecycle state buckets — drives the access decision
ACTIVE_STATES_DOCS = {
    "effective__c", "effective__v",
    "approved__c", "approved__v", "approved__sys",
    "final__c", "final__v",
    "closed__c", "closed__v", "closed__sys", "closed__qdm",
    "released__v", "released__sys",
}
DRAFT_STATES_DOCS = {
    "draft__v", "draft__c", "draft__sys",
    "in_review__v", "in_review__c",
    "initial__c", "initial__v",
    "in_approval__v",
}

ACTIVE_STATES_OBJECTS = {
    "closed_state__v", "closed__v", "closed__c", "closed__qdm",
    "approved_state__v", "approved__v", "approved__c",
    "effective__v", "effective__c",
    "implemented_state__v", "implemented__v",
    "verified_state__v",
}

USDM_BLUE = "1F4E79"
GREEN = "C6EFCE"; GREEN_TXT = "006100"
RED = "FFC7CE"; RED_TXT = "9C0006"
YELLOW = "FFEB9C"; YELLOW_TXT = "9C5700"


# ------------------------------------------------------------------
# Vault helpers
# ------------------------------------------------------------------

def vault_auth():
    if not VAULT_USER or not VAULT_PASS:
        sys.exit("VAULT_USERNAME and VAULT_PASSWORD required in .env")
    resp = requests.post(
        f"https://{VAULT_DNS}/api/{VAULT_API}/auth",
        data={"username": VAULT_USER, "password": VAULT_PASS},
        timeout=REQUEST_TIMEOUT,
    )
    resp.raise_for_status()
    sid = resp.json().get("sessionId")
    if not sid:
        sys.exit(f"Auth failed: {resp.json()}")
    return sid


def vget(session, path, params=None):
    return session.get(
        f"https://{VAULT_DNS}/api/{VAULT_API}{path}",
        params=params or {},
        timeout=REQUEST_TIMEOUT,
    )


def vquery(session, vql):
    resp = session.post(
        f"https://{VAULT_DNS}/api/{VAULT_API}/query",
        data={"q": vql},
        timeout=REQUEST_TIMEOUT,
    )
    if resp.status_code != 200:
        return [], f"HTTP {resp.status_code}: {resp.text[:200]}"
    return resp.json().get("data", []), None


def get_users(session):
    """Return {user_id: {...}} for all users in Vault."""
    resp = vget(session, "/objects/users")
    if resp.status_code != 200:
        return {}
    users = {}
    for item in resp.json().get("users", []):
        u = item.get("user", {})
        uid = str(u.get("id", "")).strip()
        if not uid:
            continue
        users[uid] = {
            "id": uid,
            "email": (u.get("user_name__v") or "").lower(),
            "first_name": (u.get("user_first_name__v") or "").lower(),
            "last_name": (u.get("user_last_name__v") or "").lower(),
            "status": u.get("status__v", ""),
            "active": u.get("active__v", False),
            "security_profile": u.get("security_profile__v", ""),
        }
    return users


def match_admins(users):
    """Return list of admin user_ids, matching by email-contains or name."""
    admins = []
    not_found = []
    for hint in ADMIN_HINTS:
        matched = None
        for uid, u in users.items():
            email = u.get("email", "")
            ln = u.get("last_name", "")
            fn = u.get("first_name", "")
            if hint["email_contains"] in email:
                matched = uid
                break
            if ln == hint["last_name"] and fn == hint["first_name"]:
                matched = uid
                break
        if matched:
            admins.append({"id": matched, "hint": hint, "user": users[matched]})
        else:
            not_found.append(hint)
    return admins, not_found


# ------------------------------------------------------------------
# Document operations
# ------------------------------------------------------------------

def get_document_roles(session, doc_id):
    resp = vget(session, f"/objects/documents/{doc_id}/roles")
    if resp.status_code != 200:
        return [], f"HTTP {resp.status_code}"
    return resp.json().get("documentRoles", []), None


def assign_users_to_doc_role(session, doc_id, role_name, user_ids, dry_run=False):
    """POST users into a document role. Returns (success, message)."""
    if not user_ids:
        return True, "no users to add"
    if dry_run:
        return True, f"[DRY-RUN] would assign {len(user_ids)} users to {role_name}"
    
    # Vault API accepts comma-separated id list
    user_list = ",".join(str(u) for u in user_ids)
    resp = session.post(
        f"https://{VAULT_DNS}/api/{VAULT_API}/objects/documents/{doc_id}/roles/{role_name}/users",
        data={"id": user_list},
        timeout=REQUEST_TIMEOUT,
    )
    if resp.status_code == 200:
        body = resp.json()
        if body.get("responseStatus") == "SUCCESS":
            return True, f"assigned {len(user_ids)} users to {role_name}"
        return False, f"{body.get('errors')}"
    return False, f"HTTP {resp.status_code}: {resp.text[:200]}"


# ------------------------------------------------------------------
# Object operations
# ------------------------------------------------------------------

def get_object_roles(session, object_name, object_id):
    resp = vget(session, f"/vobjects/{object_name}/{object_id}/roles")
    if resp.status_code != 200:
        return [], f"HTTP {resp.status_code}"
    return resp.json().get("data", []), None


def set_object_owner(session, object_name, object_id, user_id, dry_run=False):
    """PUT owner__sys on an object record."""
    if dry_run:
        return True, f"[DRY-RUN] would set owner__sys={user_id}"
    resp = session.put(
        f"https://{VAULT_DNS}/api/{VAULT_API}/vobjects/{object_name}/{object_id}",
        data={"owner__sys": user_id},
        timeout=REQUEST_TIMEOUT,
    )
    if resp.status_code == 200 and resp.json().get("responseStatus") == "SUCCESS":
        return True, f"set owner__sys = {user_id}"
    return False, f"HTTP {resp.status_code}: {resp.text[:200]}"


def assign_users_to_object_role(session, object_name, object_id, role_name, user_ids, dry_run=False):
    """Assign users to an object record's DAC role."""
    if not user_ids:
        return True, "no users to add"
    if dry_run:
        return True, f"[DRY-RUN] would assign {len(user_ids)} users to {role_name}"
    user_list = ",".join(str(u) for u in user_ids)
    resp = session.post(
        f"https://{VAULT_DNS}/api/{VAULT_API}/vobjects/{object_name}/{object_id}/roles/{role_name}/users",
        data={"id": user_list},
        timeout=REQUEST_TIMEOUT,
    )
    if resp.status_code == 200:
        body = resp.json()
        if body.get("responseStatus") == "SUCCESS":
            return True, f"assigned {len(user_ids)} users to {role_name}"
        return False, f"{body.get('errors')}"
    return False, f"HTTP {resp.status_code}: {resp.text[:200]}"


# ------------------------------------------------------------------
# Planning + applying
# ------------------------------------------------------------------

def plan_document_changes(session, users, admins, all_active_user_ids):
    """For each doc, decide what role assignments are needed."""
    print("  Querying documents...")
    docs, err = vquery(session,
                       "SELECT id, document_number__v, name__v, type__v, subtype__v, "
                       "status__v, lifecycle__v, created_by__v FROM documents MAXROWS 200")
    if err:
        print(f"  ! {err}")
        return []
    
    admin_ids = [a["id"] for a in admins]
    plans = []
    
    for i, doc in enumerate(docs, 1):
        doc_id = doc.get("id")
        status = (doc.get("status__v") or "").lower()
        creator = str(doc.get("created_by__v") or "").strip()
        
        # Current roles
        roles, _ = get_document_roles(session, doc_id)
        current_owner_users = set()
        current_viewer_users = set()
        for r in roles or []:
            rname = r.get("name", "")
            assigned = [str(u) for u in (r.get("assignedUsers") or [])]
            default = [str(u) for u in (r.get("defaultUsers") or [])]
            all_in_role = set(assigned + default)
            if rname == "owner__v":
                current_owner_users = all_in_role
            elif rname == "viewer__v":
                current_viewer_users = all_in_role
        
        # Plan: Owner role = creator + admins (deduped)
        target_owner = set([creator] if creator else []) | set(admin_ids)
        owner_to_add = target_owner - current_owner_users
        
        # Plan: Viewer role
        if status in ACTIVE_STATES_DOCS:
            target_viewer = set(all_active_user_ids)  # all users
            access_level = "ALL_USERS"
        elif status in DRAFT_STATES_DOCS:
            target_viewer = set([creator] if creator else []) | set(admin_ids)
            access_level = "AUTHOR_ONLY"
        else:
            # Unknown status — be conservative, admin + creator only
            target_viewer = set([creator] if creator else []) | set(admin_ids)
            access_level = "UNKNOWN_STATUS"
        
        viewer_to_add = target_viewer - current_viewer_users
        
        plans.append({
            "doc_id": doc_id,
            "document_number": doc.get("document_number__v", ""),
            "name": doc.get("name__v", "")[:60],
            "status": status,
            "type": doc.get("type__v", ""),
            "creator": creator,
            "access_level": access_level,
            "current_owner_count": len(current_owner_users),
            "current_viewer_count": len(current_viewer_users),
            "owner_to_add": sorted(owner_to_add),
            "viewer_to_add": sorted(viewer_to_add),
        })
        
        if i % 10 == 0:
            print(f"    {i}/{len(docs)} planned...")
    
    return plans


def plan_object_changes(session, users, admins, all_active_user_ids):
    """For each object type, plan role/owner changes."""
    admin_ids = [a["id"] for a in admins]
    plans = []
    
    for api_name in OBJECT_TYPES:
        print(f"  Planning {api_name}...")
        rows, err = vquery(session,
                           f"SELECT id, name__v, state__v, owner__sys, created_by__v "
                           f"FROM {api_name} MAXROWS 500")
        if err:
            print(f"    ! {err}")
            continue
        if not rows:
            continue
        
        for rec in rows:
            rid = rec.get("id")
            if not rid:
                continue
            
            state = (rec.get("state__v") or "").lower()
            owner_id = str(rec.get("owner__sys") or "").strip()
            creator = str(rec.get("created_by__v") or "").strip()
            
            # Plan owner__sys
            new_owner_id = None
            if not owner_id or owner_id not in users:
                # Set owner to creator (or first admin if creator is missing)
                new_owner_id = creator if creator and creator in users else (admin_ids[0] if admin_ids else None)
            
            # Plan DAC role assignments
            obj_roles, _ = get_object_roles(session, api_name, rid)
            current_viewer_users = set()
            for r in obj_roles or []:
                rname = r.get("name", "")
                if rname in ("viewer__v", "viewer_role__v"):
                    current_viewer_users = {str(u) for u in (r.get("users") or [])}
                    break
            
            # Decide who should be in viewer role
            if state in ACTIVE_STATES_OBJECTS:
                target_viewer = set(all_active_user_ids)
                access_level = "ALL_USERS"
            else:
                # Open/draft state — admins + owner only
                target_viewer = set(admin_ids) | set([owner_id or creator])
                access_level = "ADMINS_AND_OWNER"
            
            viewer_to_add = target_viewer - current_viewer_users
            
            plans.append({
                "object_type": api_name,
                "record_id": rid,
                "name": (rec.get("name__v") or "")[:60],
                "state": state,
                "current_owner": owner_id,
                "new_owner": new_owner_id or "",
                "creator": creator,
                "access_level": access_level,
                "current_viewer_count": len(current_viewer_users),
                "viewer_to_add": sorted(viewer_to_add),
            })
    
    return plans


def apply_document_plan(session, plans, dry_run=False):
    """Apply doc role changes."""
    changes = []
    now = datetime.datetime.now().isoformat()
    
    for p in plans:
        # Owner role
        if p["owner_to_add"]:
            ok, msg = assign_users_to_doc_role(session, p["doc_id"], "owner__v",
                                                p["owner_to_add"], dry_run=dry_run)
            changes.append({
                "timestamp": now,
                "target_type": "Document",
                "target_id": p["doc_id"],
                "target_label": f"{p['document_number']}: {p['name']}",
                "action": "assign owner__v",
                "users_added": ",".join(p["owner_to_add"]),
                "user_count": len(p["owner_to_add"]),
                "success": ok,
                "message": msg,
            })
            if not dry_run:
                time.sleep(0.3)
        
        # Viewer role
        if p["viewer_to_add"]:
            ok, msg = assign_users_to_doc_role(session, p["doc_id"], "viewer__v",
                                                p["viewer_to_add"], dry_run=dry_run)
            changes.append({
                "timestamp": now,
                "target_type": "Document",
                "target_id": p["doc_id"],
                "target_label": f"{p['document_number']}: {p['name']} ({p['access_level']})",
                "action": "assign viewer__v",
                "users_added": ",".join(p["viewer_to_add"]),
                "user_count": len(p["viewer_to_add"]),
                "success": ok,
                "message": msg,
            })
            if not dry_run:
                time.sleep(0.3)
    
    return changes


def apply_object_plan(session, plans, dry_run=False):
    """Apply object owner + DAC changes."""
    changes = []
    now = datetime.datetime.now().isoformat()
    
    for p in plans:
        # Set owner if needed
        if p["new_owner"]:
            ok, msg = set_object_owner(session, p["object_type"], p["record_id"],
                                       p["new_owner"], dry_run=dry_run)
            changes.append({
                "timestamp": now,
                "target_type": p["object_type"],
                "target_id": p["record_id"],
                "target_label": p["name"],
                "action": "set owner__sys",
                "users_added": p["new_owner"],
                "user_count": 1,
                "success": ok,
                "message": msg,
            })
            if not dry_run:
                time.sleep(0.3)
        
        # Viewer role (DAC)
        if p["viewer_to_add"]:
            ok, msg = assign_users_to_object_role(session, p["object_type"], p["record_id"],
                                                    "viewer__v", p["viewer_to_add"],
                                                    dry_run=dry_run)
            changes.append({
                "timestamp": now,
                "target_type": p["object_type"],
                "target_id": p["record_id"],
                "target_label": f"{p['name']} ({p['access_level']})",
                "action": "assign viewer__v",
                "users_added": ",".join(p["viewer_to_add"]),
                "user_count": len(p["viewer_to_add"]),
                "success": ok,
                "message": msg,
            })
            if not dry_run:
                time.sleep(0.3)
    
    return changes


# ------------------------------------------------------------------
# Excel report
# ------------------------------------------------------------------

def style_header(ws, row=1):
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill(start_color=USDM_BLUE, end_color=USDM_BLUE, fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def autosize(ws, max_w=60):
    for col in ws.columns:
        try:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        except ValueError:
            max_len = 0
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 8), max_w)


def color_cell(ws, row, col, val, fg, bg):
    c = ws.cell(row=row, column=col)
    c.value = val
    c.font = Font(color=fg, bold=True)
    c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")


def write_report(output, users, admins, admins_not_found, doc_plans, obj_plans, doc_changes, obj_changes, dry_run, applied):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    # Summary
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Vault GxP Security Apply"
    ws["A1"].font = Font(bold=True, size=16, color=USDM_BLUE)
    ws["A2"] = f"Vault: {VAULT_DNS}"
    ws["A3"] = f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A4"] = "Mode: " + ("DRY-RUN (no changes)" if dry_run else ("APPLIED" if applied else "AUDIT-ONLY"))
    
    ws["A6"] = "ADMIN USERS (access to everything)"
    ws["A6"].font = Font(bold=True, color="FFFFFF")
    ws["A6"].fill = PatternFill(start_color=USDM_BLUE, end_color=USDM_BLUE, fill_type="solid")
    ws.merge_cells("A6:D6")
    
    ws.append(["Hint matched", "User ID", "Email", "Active"])
    style_header(ws, row=7)
    for a in admins:
        u = a["user"]
        ws.append([a["hint"]["email_contains"], u["id"], u["email"], u["active"]])
    
    if admins_not_found:
        for hint in admins_not_found:
            ws.append([hint["email_contains"], "NOT FOUND", "", ""])
            color_cell(ws, ws.max_row, 2, "NOT FOUND", RED_TXT, RED)
    
    next_row = ws.max_row + 2
    ws.cell(row=next_row, column=1).value = "PLAN SUMMARY"
    ws.cell(row=next_row, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=next_row, column=1).fill = PatternFill(start_color=USDM_BLUE, end_color=USDM_BLUE, fill_type="solid")
    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=4)
    
    metrics = [
        ("Active users found", sum(1 for u in users.values() if u["active"])),
        ("Documents planned", len(doc_plans)),
        ("  → docs with role changes", sum(1 for p in doc_plans if p["owner_to_add"] or p["viewer_to_add"])),
        ("Object records planned", len(obj_plans)),
        ("  → records needing owner__sys", sum(1 for p in obj_plans if p["new_owner"])),
        ("  → records with DAC changes", sum(1 for p in obj_plans if p["viewer_to_add"])),
        ("", ""),
        ("Doc API calls", len(doc_changes)),
        ("  → success", sum(1 for c in doc_changes if c["success"])),
        ("  → failed", sum(1 for c in doc_changes if not c["success"])),
        ("Object API calls", len(obj_changes)),
        ("  → success", sum(1 for c in obj_changes if c["success"])),
        ("  → failed", sum(1 for c in obj_changes if not c["success"])),
    ]
    for label, val in metrics:
        ws.append([label, val, "", ""])
    
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 15
    
    # Security Model
    ws = wb.create_sheet("Security Model")
    ws["A1"] = "Security Model Applied"
    ws["A1"].font = Font(bold=True, size=14, color=USDM_BLUE)
    
    ws.append(["", ""])
    ws.append(["Rule", "Detail"])
    style_header(ws, row=3)
    
    rules = [
        ("Admin users (full access)", "Jennell Botero, Brittany Walker, Hovsep Kirikian — assigned to owner__v on ALL docs and DAC viewer on ALL objects"),
        ("Docs: Effective / Approved / Closed", "All active Vault users get viewer__v role"),
        ("Docs: Draft / In Review", "Author (created_by) + admins only"),
        ("Objects: Closed / Approved", "All active users via DAC viewer role"),
        ("Objects: Open / Draft / In Progress", "Owner + admins only"),
        ("Objects with no owner__sys", "Set to created_by, or first admin if creator missing"),
    ]
    for r in rules:
        ws.append(r)
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 95
    
    # Document Plan
    ws = wb.create_sheet("Document Plan")
    headers_list = ["doc_id", "document_number", "name", "type", "status", "access_level",
                    "current_owner_count", "current_viewer_count",
                    "owner_to_add_count", "viewer_to_add_count"]
    ws.append(headers_list)
    style_header(ws)
    
    for p in doc_plans:
        ws.append([p["doc_id"], p["document_number"], p["name"], p["type"], p["status"],
                  p["access_level"], p["current_owner_count"], p["current_viewer_count"],
                  len(p["owner_to_add"]), len(p["viewer_to_add"])])
        row = ws.max_row
        col = headers_list.index("access_level") + 1
        if p["access_level"] == "ALL_USERS":
            color_cell(ws, row, col, "ALL_USERS", GREEN_TXT, GREEN)
        elif p["access_level"] == "AUTHOR_ONLY":
            color_cell(ws, row, col, "AUTHOR_ONLY", YELLOW_TXT, YELLOW)
        else:
            color_cell(ws, row, col, p["access_level"], RED_TXT, RED)
    
    ws.freeze_panes = "A2"
    autosize(ws, max_w=40)
    
    # Object Plan
    ws = wb.create_sheet("Object Plan")
    obj_headers = ["object_type", "record_id", "name", "state", "access_level",
                   "current_owner", "new_owner", "current_viewer_count", "viewer_to_add_count"]
    ws.append(obj_headers)
    style_header(ws)
    for p in obj_plans:
        ws.append([p["object_type"], p["record_id"], p["name"], p["state"], p["access_level"],
                  p["current_owner"], p["new_owner"], p["current_viewer_count"], len(p["viewer_to_add"])])
        row = ws.max_row
        col = obj_headers.index("access_level") + 1
        if p["access_level"] == "ALL_USERS":
            color_cell(ws, row, col, "ALL_USERS", GREEN_TXT, GREEN)
        else:
            color_cell(ws, row, col, p["access_level"], YELLOW_TXT, YELLOW)
    ws.freeze_panes = "A2"
    autosize(ws, max_w=40)
    
    # API Changes
    if doc_changes or obj_changes:
        ws = wb.create_sheet("API Changes Log")
        chg_headers = ["timestamp", "target_type", "target_id", "target_label",
                       "action", "user_count", "success", "message"]
        ws.append(chg_headers)
        style_header(ws)
        for c in doc_changes + obj_changes:
            ws.append([c.get(h, "") for h in chg_headers])
            row = ws.max_row
            scol = chg_headers.index("success") + 1
            if c["success"]:
                color_cell(ws, row, scol, "TRUE", GREEN_TXT, GREEN)
            else:
                color_cell(ws, row, scol, "FALSE", RED_TXT, RED)
        ws.freeze_panes = "A2"
        autosize(ws, max_w=60)
    
    # All users sheet
    ws = wb.create_sheet("All Users")
    ws.append(["User ID", "Email", "First Name", "Last Name", "Status", "Active", "Security Profile"])
    style_header(ws)
    for u in sorted(users.values(), key=lambda x: x["email"]):
        ws.append([u["id"], u["email"], u["first_name"], u["last_name"],
                  u["status"], u["active"], u["security_profile"]])
    ws.freeze_panes = "A2"
    autosize(ws)
    
    wb.save(output)


# ------------------------------------------------------------------
# Main
# ------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="Apply GxP security model to Vault")
    ap.add_argument("--audit-only", action="store_true", help="Plan only, no writes")
    ap.add_argument("--dry-run", action="store_true", help="Show API calls without making them")
    ap.add_argument("--apply", action="store_true", help="Actually write to Vault")
    ap.add_argument("--output", default=None)
    args = ap.parse_args()
    
    if not (args.audit_only or args.dry_run or args.apply):
        print("\nMust specify exactly one of:")
        print("  --audit-only    Plan and show what's there (read-only)")
        print("  --dry-run       Plan changes, simulate API calls, no writes")
        print("  --apply         Plan and execute changes (writes to Vault)")
        sys.exit(1)
    
    if not args.output:
        stamp = datetime.datetime.now().strftime("%Y-%m-%d_%H%M")
        args.output = f"Vault_GxP_Security_{stamp}.xlsx"
    
    print("=" * 80)
    print(f"VAULT GxP SECURITY APPLY — {VAULT_DNS}")
    mode = "AUDIT-ONLY" if args.audit_only else ("DRY-RUN" if args.dry_run else "APPLY")
    print(f"  Mode:   {mode}")
    print(f"  Output: {args.output}")
    print("=" * 80)
    
    sid = vault_auth()
    session = requests.Session()
    session.headers.update({"Authorization": sid, "Accept": "application/json"})
    print("✓ Vault auth OK\n")
    
    print("[1/5] Fetching users...")
    users = get_users(session)
    print(f"  ✓ {len(users)} users")
    
    print("\n[2/5] Identifying admins...")
    admins, not_found = match_admins(users)
    for a in admins:
        u = a["user"]
        print(f"  ✓ {a['hint']['first_name'].title()} {a['hint']['last_name'].title()} → {u['email']} (id={u['id']})")
    for nf in not_found:
        print(f"  ✗ Not found: {nf['first_name'].title()} {nf['last_name'].title()} (hint: {nf['email_contains']})")
    
    if not admins:
        sys.exit("ERROR: No admin users matched. Edit ADMIN_HINTS at top of script.")
    
    active_users = [uid for uid, u in users.items() if u["active"]]
    print(f"\n  Active users available for ALL_USERS access: {len(active_users)}")
    
    print("\n[3/5] Planning document changes...")
    doc_plans = plan_document_changes(session, users, admins, active_users)
    print(f"  ✓ {len(doc_plans)} documents planned")
    
    print("\n[4/5] Planning object changes...")
    obj_plans = plan_object_changes(session, users, admins, active_users)
    print(f"  ✓ {len(obj_plans)} object records planned")
    
    doc_changes = []
    obj_changes = []
    
    if args.dry_run or args.apply:
        print(f"\n[5/5] {'Simulating' if args.dry_run else 'Applying'} changes...")
        doc_changes = apply_document_plan(session, doc_plans, dry_run=args.dry_run)
        obj_changes = apply_object_plan(session, obj_plans, dry_run=args.dry_run)
        
        ok = sum(1 for c in doc_changes + obj_changes if c["success"])
        bad = len(doc_changes) + len(obj_changes) - ok
        print(f"  Total API calls: {len(doc_changes) + len(obj_changes)}")
        print(f"    Success: {ok}")
        print(f"    Failed:  {bad}")
        
        if args.apply and (doc_changes or obj_changes):
            with open("changes_applied.csv", "w", newline="") as f:
                all_changes = doc_changes + obj_changes
                w = csv.DictWriter(f, fieldnames=all_changes[0].keys())
                w.writeheader()
                w.writerows(all_changes)
            print(f"  Changes log: changes_applied.csv")
    
    print(f"\nBuilding report: {args.output}")
    write_report(args.output, users, admins, not_found, doc_plans, obj_plans,
                doc_changes, obj_changes, args.dry_run, args.apply)
    print(f"✓ Report saved")
    
    print("\n" + "=" * 80)
    print("DONE")
    print("=" * 80)
    
    if args.apply:
        print("\nNext steps:")
        print("  1. python vault_to_glean_sync_v2.py     (re-push to Glean with new ACLs)")
        print("  2. python glean_monitor.py --user-email jbotero@usdm.com")
        print("     (verify you can now see content)")


if __name__ == "__main__":
    main()
