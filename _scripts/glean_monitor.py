#!/usr/bin/env python3
"""
Comprehensive Glean monitoring report for veevavaultquality datasource.

Runs all troubleshooting checks in one pass and produces a single Excel workbook
with multiple sheets covering: document counts, datasource config, run history,
per-doc indexing status, ACL audit, silent docs, per-user access checks, and
recommended actions.

Usage:
    export GLEAN_API_URL=https://usdm-be.glean.com
    export GLEAN_INDEXING_API_TOKEN=<your token>
    export DATASOURCE_NAME=veevavaultquality

    # Full report (all sheets, no user-access check)
    python glean_monitor.py

    # Full report with per-user access check (adds ~3.5 min)
    python glean_monitor.py --user-email jbotero@usdm.com

    # Multiple users
    python glean_monitor.py --user-email jbotero@usdm.com --user-email hkirikian@usdm.com

    # Skip the slow per-doc check (just config + counts + run history)
    python glean_monitor.py --no-batch-check

Outputs:
    Glean_Monitor_Report_YYYY-MM-DD_HHMM.xlsx
"""

import argparse
import datetime
import os
import sys
import time
from collections import Counter
from pathlib import Path

import pandas as pd
import requests
from dotenv import load_dotenv
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

load_dotenv()

GLEAN_URL = os.getenv("GLEAN_API_URL", "").rstrip("/")
GLEAN_TOKEN = os.getenv("GLEAN_INDEXING_API_TOKEN") or os.getenv("GLEAN_API_TOKEN")
DATASOURCE = os.getenv("DATASOURCE_NAME", "veevavaultquality")
RATE_SLEEP = float(os.getenv("DEBUG_RATE_SLEEP", "1.1"))
REQUEST_TIMEOUT = int(os.getenv("REQUEST_TIMEOUT", "60"))


def require_env():
    if not GLEAN_URL:
        sys.exit("ERROR: GLEAN_API_URL not set")
    if not GLEAN_TOKEN:
        sys.exit("ERROR: GLEAN_INDEXING_API_TOKEN not set")


def headers():
    return {
        "Authorization": f"Bearer {GLEAN_TOKEN}",
        "Content-Type": "application/json",
    }


# ------------------------------------------------------------------
# Glean API wrappers
# ------------------------------------------------------------------

def get_document_count(session):
    resp = session.post(
        f"{GLEAN_URL}/api/index/v1/getdocumentcount",
        headers=headers(),
        json={"datasource": DATASOURCE},
        timeout=REQUEST_TIMEOUT,
    )
    if resp.status_code != 200:
        return None, f"HTTP {resp.status_code}: {resp.text[:300]}"
    return resp.json(), None


def get_datasource_config(session):
    resp = session.post(
        f"{GLEAN_URL}/api/index/v1/getdatasourceconfig",
        headers=headers(),
        json={"datasource": DATASOURCE},
        timeout=REQUEST_TIMEOUT,
    )
    if resp.status_code != 200:
        return None, f"HTTP {resp.status_code}: {resp.text[:300]}"
    return resp.json(), None


def get_debug_status(session):
    resp = session.post(
        f"{GLEAN_URL}/api/index/v1/debug/{DATASOURCE}/status",
        headers=headers(),
        json={},
        timeout=REQUEST_TIMEOUT,
    )
    if resp.status_code != 200:
        return None, f"HTTP {resp.status_code}: {resp.text[:300]}"
    return resp.json(), None


def get_debug_document(session, object_type, doc_id):
    resp = session.post(
        f"{GLEAN_URL}/api/index/v1/debug/{DATASOURCE}/document",
        headers=headers(),
        json={"objectType": object_type, "docId": doc_id},
        timeout=REQUEST_TIMEOUT,
    )
    if resp.status_code == 404:
        return {"_status": "NOT_FOUND"}, None
    if resp.status_code == 429:
        return {"_status": "RATE_LIMITED"}, "Rate limited"
    if resp.status_code != 200:
        return None, f"HTTP {resp.status_code}: {resp.text[:200]}"
    data = resp.json()
    data["_status"] = "INDEXED"
    return data, None


def check_document_access(session, object_type, doc_id, user_email):
    resp = session.post(
        f"{GLEAN_URL}/api/index/v1/checkdocumentaccess",
        headers=headers(),
        json={
            "datasource": DATASOURCE,
            "objectType": object_type,
            "docId": doc_id,
            "userEmail": user_email,
        },
        timeout=REQUEST_TIMEOUT,
    )
    if resp.status_code == 200:
        return resp.json().get("hasAccess", False), None
    if resp.status_code == 429:
        return None, "RATE_LIMITED"
    return None, f"HTTP {resp.status_code}: {resp.text[:120]}"


# ------------------------------------------------------------------
# Parsers
# ------------------------------------------------------------------

def parse_debug_doc(raw):
    out = {}
    if not isinstance(raw, dict):
        return out

    out["upload_status"] = raw.get("uploadStatus", "")
    out["indexing_status"] = raw.get("indexingStatus", "")

    for src, dst in [("lastUploadedAt", "last_uploaded"),
                     ("lastIndexedAt", "last_indexed")]:
        ts = raw.get(src)
        if isinstance(ts, (int, float)) and ts > 0:
            try:
                out[dst] = datetime.datetime.utcfromtimestamp(ts).isoformat() + "Z"
            except Exception:
                out[dst] = str(ts)
        else:
            out[dst] = ""

    perms = raw.get("uploadedPermissions") or {}
    if isinstance(perms, dict):
        out["allow_anonymous"] = bool(perms.get("allowAnonymousAccess", False))
        out["allow_all_ds_users"] = bool(perms.get("allowAllDatasourceUsersAccess", False))

        allowed_users = perms.get("allowedUsers") or []
        out["allowed_users_count"] = len(allowed_users) if isinstance(allowed_users, list) else 0
        sample_u = []
        for u in (allowed_users[:5] if isinstance(allowed_users, list) else []):
            if isinstance(u, dict):
                sample_u.append(u.get("email") or u.get("datasourceUserId") or "?")
            else:
                sample_u.append(str(u))
        out["allowed_users_sample"] = "; ".join(sample_u)

        allowed_groups = perms.get("allowedGroups") or []
        out["allowed_groups_count"] = len(allowed_groups) if isinstance(allowed_groups, list) else 0
        sample_g = []
        for g in (allowed_groups[:5] if isinstance(allowed_groups, list) else []):
            if isinstance(g, dict):
                sample_g.append(g.get("name") or g.get("groupId") or "?")
            else:
                sample_g.append(str(g))
        out["allowed_groups_sample"] = "; ".join(sample_g)

    return out


def load_batch_file(path):
    items = []
    with open(path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "," in line:
                parts = [p.strip() for p in line.split(",", 1)]
                items.append((parts[0], parts[1]))
            else:
                items.append(("Document", line))
    return items


# ------------------------------------------------------------------
# Excel styling
# ------------------------------------------------------------------

USDM_BLUE = "1F4E79"
GREEN = "C6EFCE"
GREEN_TXT = "006100"
RED = "FFC7CE"
RED_TXT = "9C0006"
YELLOW = "FFEB9C"
YELLOW_TXT = "9C5700"
GRAY = "E7E6E6"


def style_header_row(ws, row=1):
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill(start_color=USDM_BLUE, end_color=USDM_BLUE, fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def autosize(ws, min_w=8, max_w=60):
    for col in ws.columns:
        col_letter = col[0].column_letter
        try:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        except ValueError:
            max_len = 0
        ws.column_dimensions[col_letter].width = max(min_w, min(max_w, max_len + 2))


def color_cell(ws, row, col, text, fg, bg):
    cell = ws.cell(row=row, column=col)
    cell.value = text
    cell.font = Font(color=fg, bold=True, name="Arial", size=10)
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")


# ------------------------------------------------------------------
# Sheet builders
# ------------------------------------------------------------------

def sheet_summary(wb, ctx):
    ws = wb.create_sheet("Summary", 0)
    ws["A1"] = "Glean Monitoring Report"
    ws["A1"].font = Font(bold=True, size=16, color=USDM_BLUE)
    ws["A2"] = f"Datasource: {DATASOURCE}"
    ws["A3"] = f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A4"] = f"Glean URL: {GLEAN_URL}"

    # Key metrics block
    ws["A6"] = "KEY METRICS"
    ws["A6"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A6"].fill = PatternFill(start_color=USDM_BLUE, end_color=USDM_BLUE, fill_type="solid")
    ws.merge_cells("A6:C6")

    metrics = [
        ("Total docs in datasource", ctx.get("doc_count", "?")),
        ("Items expected from Vault", ctx.get("expected_total", "?")),
        ("Items confirmed INDEXED", ctx.get("indexed_count", "?")),
        ("Items NOT_FOUND", ctx.get("not_found_count", "?")),
        ("Items with ERROR", ctx.get("error_count", "?")),
        ("Indexing success rate", f"{ctx.get('indexed_pct', 0):.0f}%"),
        ("", ""),
        ("ACL — visible to someone", ctx.get("visible_count", "?")),
        ("ACL — silent (no permission)", ctx.get("silent_count", "?")),
        ("ACL — allowAnonymous", ctx.get("anon_count", 0)),
        ("ACL — allowAllDatasourceUsers", ctx.get("all_ds_count", 0)),
        ("ACL — specific allowedUsers", ctx.get("user_acl_count", 0)),
        ("ACL — specific allowedGroups", ctx.get("group_acl_count", 0)),
    ]
    for i, (label, value) in enumerate(metrics, start=7):
        ws.cell(row=i, column=1).value = label
        ws.cell(row=i, column=2).value = value
        if label.startswith("ACL — silent") and isinstance(value, int) and value > 0:
            ws.cell(row=i, column=2).fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
        elif "ERROR" in str(label) and isinstance(value, int) and value > 0:
            ws.cell(row=i, column=2).fill = PatternFill(start_color=RED, end_color=RED, fill_type="solid")

    # Status block
    next_row = 7 + len(metrics) + 2
    ws.cell(row=next_row, column=1).value = "STATUS"
    ws.cell(row=next_row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=next_row, column=1).fill = PatternFill(start_color=USDM_BLUE, end_color=USDM_BLUE, fill_type="solid")
    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=3)

    status_lines = ctx.get("status_lines", [])
    for i, line in enumerate(status_lines, start=next_row + 1):
        ws.cell(row=i, column=1).value = line.get("label", "")
        color_cell(ws, i, 2, line.get("verdict", ""), line.get("fg", "000000"), line.get("bg", GRAY))
        ws.cell(row=i, column=3).value = line.get("detail", "")

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 90


def sheet_config(wb, config):
    ws = wb.create_sheet("Datasource Config")
    ws.append(["Field", "Value"])
    style_header_row(ws)

    if not config:
        ws.append(["ERROR", "Could not fetch config"])
        ws.cell(row=2, column=2).fill = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
        return

    rows = [
        ("name", config.get("name", "")),
        ("displayName", config.get("displayName", "")),
        ("homeUrl", config.get("homeUrl", "")),
        ("datasourceCategory", config.get("datasourceCategory", "")),
        ("urlRegex", config.get("urlRegex", "")),
        ("isUserReferencedByEmail", str(config.get("isUserReferencedByEmail", ""))),
        ("isOnPrem", str(config.get("isOnPrem", ""))),
        ("isTestDatasource", str(config.get("isTestDatasource", ""))),
        ("isEntityDatasource", str(config.get("isEntityDatasource", ""))),
        ("object types configured", len(config.get("objectDefinitions", []))),
        ("", ""),
        ("OBJECT TYPE NAME", "CATEGORY"),
    ]
    for r in rows:
        ws.append(r)
    # Mark the second header
    ws.cell(row=len(rows) + 1, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=len(rows) + 1, column=2).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=len(rows) + 1, column=1).fill = PatternFill(start_color=USDM_BLUE, end_color=USDM_BLUE, fill_type="solid")
    ws.cell(row=len(rows) + 1, column=2).fill = PatternFill(start_color=USDM_BLUE, end_color=USDM_BLUE, fill_type="solid")

    for od in config.get("objectDefinitions", []):
        ws.append([od.get("name", ""), od.get("docCategory", "")])

    autosize(ws)


def sheet_run_history(wb, status):
    ws = wb.create_sheet("Run History")
    if not status:
        ws.append(["ERROR", "Could not fetch run history"])
        return

    bulk = status.get("bulkUploadHistory", []) or []
    proc = status.get("processingHistory", []) or []

    ws.append(["LAST 10 BULK UPLOADS"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill(start_color=USDM_BLUE, end_color=USDM_BLUE, fill_type="solid")
    ws.merge_cells("A1:C1")

    ws.append(["#", "Datetime", "Upload ID"])
    style_header_row(ws, row=2)
    if not bulk:
        ws.append(["", "No bulk upload history found", ""])
    else:
        for i, e in enumerate(bulk[:10], 1):
            ws.append([i, e.get("datetime", ""), e.get("uploadId", "")])

    blank_row = ws.max_row + 2
    ws.cell(row=blank_row, column=1).value = "LAST 10 PROCESSING RUNS"
    ws.cell(row=blank_row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=blank_row, column=1).fill = PatternFill(start_color=USDM_BLUE, end_color=USDM_BLUE, fill_type="solid")
    ws.merge_cells(start_row=blank_row, start_column=1, end_row=blank_row, end_column=3)

    ws.append(["#", "Datetime", ""])
    style_header_row(ws, row=blank_row + 1)
    if not proc:
        ws.append(["", "No processing history found", ""])
    else:
        for i, e in enumerate(proc[:10], 1):
            ws.append([i, e.get("datetime", ""), ""])

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 50


def sheet_per_doc(wb, results):
    ws = wb.create_sheet("Per-Doc Indexing")
    if not results:
        ws.append(["No results"])
        return

    headers_list = [
        "object_type", "doc_id", "status",
        "upload_status", "indexing_status",
        "last_uploaded", "last_indexed",
        "allow_anonymous", "allow_all_ds_users",
        "allowed_users_count", "allowed_users_sample",
        "allowed_groups_count", "allowed_groups_sample",
        "detail",
    ]
    ws.append(headers_list)
    style_header_row(ws)

    for r in results:
        row = [r.get(h, "") for h in headers_list]
        ws.append(row)
        # Color status column
        status = r.get("status", "")
        row_num = ws.max_row
        col_idx = headers_list.index("status") + 1
        if status == "INDEXED":
            color_cell(ws, row_num, col_idx, status, GREEN_TXT, GREEN)
        elif status == "NOT_FOUND":
            color_cell(ws, row_num, col_idx, status, YELLOW_TXT, YELLOW)
        elif status == "ERROR":
            color_cell(ws, row_num, col_idx, status, RED_TXT, RED)

    ws.freeze_panes = "A2"
    autosize(ws, max_w=40)


def sheet_coverage_by_type(wb, results):
    ws = wb.create_sheet("Coverage by Object Type")
    ws.append(["Object Type", "INDEXED", "NOT_FOUND", "ERROR", "Total"])
    style_header_row(ws)

    by_type = {}
    for r in results:
        t = r.get("object_type", "")
        by_type.setdefault(t, {"INDEXED": 0, "NOT_FOUND": 0, "ERROR": 0})
        by_type[t][r.get("status", "")] = by_type[t].get(r.get("status", ""), 0) + 1

    for t in sorted(by_type.keys()):
        c = by_type[t]
        total = sum(c.values())
        ws.append([t, c.get("INDEXED", 0), c.get("NOT_FOUND", 0), c.get("ERROR", 0), total])

    autosize(ws)


def sheet_acl_summary(wb, results, ctx):
    ws = wb.create_sheet("ACL Analysis")
    indexed = [r for r in results if r.get("status") == "INDEXED"]
    if not indexed:
        ws.append(["No INDEXED rows"])
        return

    # High-level breakdown
    ws.append(["ACL PATTERN", "COUNT", "VISIBILITY"])
    style_header_row(ws)

    anon = sum(1 for r in indexed if r.get("allow_anonymous"))
    all_ds = sum(1 for r in indexed if r.get("allow_all_ds_users"))
    user_only = sum(1 for r in indexed if (r.get("allowed_users_count", 0) or 0) > 0
                    and not r.get("allow_anonymous") and not r.get("allow_all_ds_users"))
    group_only = sum(1 for r in indexed if (r.get("allowed_groups_count", 0) or 0) > 0
                     and (r.get("allowed_users_count", 0) or 0) == 0
                     and not r.get("allow_anonymous") and not r.get("allow_all_ds_users"))
    silent = sum(1 for r in indexed if (r.get("allowed_users_count", 0) or 0) == 0
                 and (r.get("allowed_groups_count", 0) or 0) == 0
                 and not r.get("allow_anonymous") and not r.get("allow_all_ds_users"))

    rows = [
        ("allowAnonymousAccess=True", anon, "Anyone (rare for biopharma)"),
        ("allowAllDatasourceUsersAccess=True", all_ds, "Any user with datasource access"),
        ("Specific allowedUsers only", user_only, "Listed users only"),
        ("Specific allowedGroups only", group_only, "Listed group members only"),
        ("No ACLs at all (silent)", silent, "Invisible to all users in Glean search"),
    ]
    for r in rows:
        ws.append(r)
        if r[0].startswith("No ACLs") and r[1] > 0:
            ws.cell(row=ws.max_row, column=2).fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")

    # Silent docs by object type
    blank = ws.max_row + 2
    ws.cell(row=blank, column=1).value = "SILENT DOCS BY OBJECT TYPE"
    ws.cell(row=blank, column=1).font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=blank, column=1).fill = PatternFill(start_color=USDM_BLUE, end_color=USDM_BLUE, fill_type="solid")
    ws.merge_cells(start_row=blank, start_column=1, end_row=blank, end_column=4)

    ws.append(["Object Type", "Silent", "Visible", "Total"])
    style_header_row(ws, row=ws.max_row)

    by_type = {}
    for r in indexed:
        t = r.get("object_type", "")
        is_silent = ((r.get("allowed_users_count", 0) or 0) == 0
                     and (r.get("allowed_groups_count", 0) or 0) == 0
                     and not r.get("allow_anonymous")
                     and not r.get("allow_all_ds_users"))
        by_type.setdefault(t, {"silent": 0, "total": 0})
        by_type[t]["total"] += 1
        if is_silent:
            by_type[t]["silent"] += 1

    for t in sorted(by_type.keys(), key=lambda x: -by_type[x]["silent"]):
        c = by_type[t]
        ws.append([t, c["silent"], c["total"] - c["silent"], c["total"]])

    autosize(ws)


def sheet_silent_docs(wb, results):
    indexed = [r for r in results if r.get("status") == "INDEXED"]
    silent = [r for r in indexed if (r.get("allowed_users_count", 0) or 0) == 0
              and (r.get("allowed_groups_count", 0) or 0) == 0
              and not r.get("allow_anonymous")
              and not r.get("allow_all_ds_users")]
    if not silent:
        return

    ws = wb.create_sheet("Silent Docs (Zero ACL)")
    headers_list = ["object_type", "doc_id", "upload_status", "indexing_status", "last_indexed"]
    ws.append(headers_list)
    style_header_row(ws)
    for r in silent:
        ws.append([r.get(h, "") for h in headers_list])
    ws.freeze_panes = "A2"
    autosize(ws)


def sheet_user_access(wb, user_email, items, session):
    ws = wb.create_sheet(f"Access - {user_email[:20]}")
    ws.append(["Object Type", "Doc ID", "Has Access", "Detail"])
    style_header_row(ws)

    has_count = 0
    no_count = 0
    err_count = 0

    print(f"  Checking {user_email} against {len(items)} items...")
    for i, (obj_type, doc_id) in enumerate(items, 1):
        has_access, err = check_document_access(session, obj_type, doc_id, user_email)
        if err == "RATE_LIMITED":
            time.sleep(3.0)
            has_access, err = check_document_access(session, obj_type, doc_id, user_email)

        if err:
            ws.append([obj_type, doc_id, "ERROR", err])
            color_cell(ws, ws.max_row, 3, "ERROR", RED_TXT, RED)
            err_count += 1
        elif has_access:
            ws.append([obj_type, doc_id, "yes", ""])
            color_cell(ws, ws.max_row, 3, "yes", GREEN_TXT, GREEN)
            has_count += 1
        else:
            ws.append([obj_type, doc_id, "no", ""])
            color_cell(ws, ws.max_row, 3, "no", YELLOW_TXT, YELLOW)
            no_count += 1

        if i % 30 == 0:
            print(f"    {i}/{len(items)}: yes={has_count}, no={no_count}, err={err_count}")
        time.sleep(RATE_SLEEP)

    ws.freeze_panes = "A2"
    autosize(ws)
    return {"yes": has_count, "no": no_count, "err": err_count}


def sheet_actions(wb, ctx):
    ws = wb.create_sheet("Recommended Actions")
    ws.append(["#", "Priority", "Action", "Why"])
    style_header_row(ws)

    actions = []
    n = 1

    if ctx.get("error_count", 0) > 0:
        actions.append((n, "HIGH",
                       "Investigate ERROR items (see Per-Doc Indexing sheet, filter status=ERROR)",
                       f"{ctx['error_count']} items returned HTTP errors. If consistent, file Glean support ticket."))
        n += 1

    if ctx.get("not_found_count", 0) > 0:
        actions.append((n, "HIGH",
                       "Re-push NOT_FOUND items via glean_push_one.py",
                       f"{ctx['not_found_count']} items expected in Vault but not in Glean — sync gap."))
        n += 1

    silent = ctx.get("silent_count", 0)
    if silent > 0:
        actions.append((n, "MEDIUM",
                       "Review Silent Docs sheet (strict Vault parity is intentional)",
                       f"{silent} docs are in Glean but invisible because Vault returned no ACL. This is expected if Vault roles haven't been assigned for these object types. Confirm with QA stakeholders that these object types (e.g. Severity, RiskLevel) are SUPPOSED to be system-data and not user-visible."))
        n += 1

    if ctx.get("doc_count", 0) > ctx.get("expected_total", 0) + 10:
        diff = ctx["doc_count"] - ctx["expected_total"]
        actions.append((n, "MEDIUM",
                       f"Datasource has {diff} extra docs vs current Vault state",
                       "Could be stale items from previous syncs. Consider running a clean re-sync with bulkindexdocuments to replace the index, or selectively delete stale items via /deletedocument."))
        n += 1

    if not ctx.get("bulk_history"):
        actions.append((n, "LOW",
                       "Verify datasource is enabled for search in Glean Admin Console",
                       "Missing bulk upload history can indicate per-doc indexdocument calls (not bulk) or that processing hasn't run yet. If docs are in index but not searchable, the datasource may not be enabled for search yet — check app.glean.com/admin."))
        n += 1

    if not actions:
        actions.append((n, "NONE",
                       "All checks passed",
                       "No critical issues detected. Datasource is healthy."))

    for a in actions:
        ws.append(a)
        priority_col = ws.max_row
        if a[1] == "HIGH":
            color_cell(ws, priority_col, 2, "HIGH", RED_TXT, RED)
        elif a[1] == "MEDIUM":
            color_cell(ws, priority_col, 2, "MEDIUM", YELLOW_TXT, YELLOW)
        elif a[1] == "LOW":
            color_cell(ws, priority_col, 2, "LOW", "1F4E79", "DAE3F3")
        elif a[1] == "NONE":
            color_cell(ws, priority_col, 2, "OK", GREEN_TXT, GREEN)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 80


# ------------------------------------------------------------------
# Main
# ------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="Comprehensive Glean monitoring report")
    ap.add_argument("--batch-file", default="expected_docs.txt",
                    help="Path to expected_docs.txt (default: expected_docs.txt)")
    ap.add_argument("--no-batch-check", action="store_true",
                    help="Skip per-doc indexing check (much faster)")
    ap.add_argument("--user-email", action="append", default=[],
                    help="Email(s) to run per-user access check. Repeatable.")
    ap.add_argument("--output", default=None,
                    help="Output xlsx path (default: Glean_Monitor_Report_YYYY-MM-DD_HHMM.xlsx)")
    args = ap.parse_args()

    require_env()

    if not args.output:
        stamp = datetime.datetime.now().strftime("%Y-%m-%d_%H%M")
        args.output = f"Glean_Monitor_Report_{stamp}.xlsx"

    session = requests.Session()
    wb_ctx = {}

    print("=" * 80)
    print(f"GLEAN MONITORING — {DATASOURCE}")
    print(f"  Output: {args.output}")
    print("=" * 80)

    from openpyxl import Workbook
    wb = Workbook()
    # Remove the default sheet — we'll add Summary at the front
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ---------- 1. Document count ----------
    print("\n[1] Fetching document count...")
    counts, err = get_document_count(session)
    if err:
        print(f"  ! {err}")
        wb_ctx["doc_count"] = "ERROR"
    else:
        wb_ctx["doc_count"] = counts.get("documentCount", 0)
        print(f"  ✓ {wb_ctx['doc_count']:,} docs in datasource")

    # ---------- 2. Config ----------
    print("\n[2] Fetching datasource config...")
    config, err = get_datasource_config(session)
    if err:
        print(f"  ! {err}")
    else:
        print(f"  ✓ {len(config.get('objectDefinitions', []))} object types configured")
    sheet_config(wb, config)

    # ---------- 3. Run history ----------
    print("\n[3] Fetching run history...")
    status, err = get_debug_status(session)
    if err:
        print(f"  ! {err}")
    else:
        bulk = status.get("bulkUploadHistory", []) or []
        proc = status.get("processingHistory", []) or []
        wb_ctx["bulk_history"] = bool(bulk)
        wb_ctx["proc_history"] = bool(proc)
        print(f"  ✓ {len(bulk)} bulk uploads, {len(proc)} processing runs in history")
    sheet_run_history(wb, status)

    # ---------- 4. Per-doc indexing check ----------
    results = []
    if not args.no_batch_check:
        if not Path(args.batch_file).exists():
            print(f"\n[4] WARNING: {args.batch_file} not found. Skipping per-doc check.")
            print(f"    Run build_expected_docs.py first to generate it.")
        else:
            items = load_batch_file(args.batch_file)
            wb_ctx["expected_total"] = len(items)
            print(f"\n[4] Per-doc indexing check ({len(items)} items, ~{len(items)*RATE_SLEEP:.0f}s)")

            for i, (obj_type, doc_id) in enumerate(items, 1):
                raw, e = get_debug_document(session, obj_type, doc_id)
                if e and "Rate limited" in e:
                    time.sleep(3.0)
                    raw, e = get_debug_document(session, obj_type, doc_id)
                row = {"object_type": obj_type, "doc_id": doc_id}
                if e:
                    row["status"] = "ERROR"
                    row["detail"] = e
                elif raw and raw.get("_status") == "NOT_FOUND":
                    row["status"] = "NOT_FOUND"
                    row["detail"] = "Not in Glean index"
                elif raw and raw.get("_status") == "INDEXED":
                    row["status"] = "INDEXED"
                    row.update(parse_debug_doc(raw))
                else:
                    row["status"] = "UNKNOWN"
                results.append(row)
                if i % 30 == 0:
                    sc = Counter(r["status"] for r in results)
                    print(f"    {i}/{len(items)}: " + ", ".join(f"{s}={c}" for s, c in sc.items()))
                time.sleep(RATE_SLEEP)

            # Compute ctx
            sc = Counter(r["status"] for r in results)
            wb_ctx["indexed_count"] = sc.get("INDEXED", 0)
            wb_ctx["not_found_count"] = sc.get("NOT_FOUND", 0)
            wb_ctx["error_count"] = sc.get("ERROR", 0)
            if wb_ctx["expected_total"] > 0:
                wb_ctx["indexed_pct"] = wb_ctx["indexed_count"] / wb_ctx["expected_total"] * 100

            indexed = [r for r in results if r["status"] == "INDEXED"]
            wb_ctx["anon_count"] = sum(1 for r in indexed if r.get("allow_anonymous"))
            wb_ctx["all_ds_count"] = sum(1 for r in indexed if r.get("allow_all_ds_users"))
            wb_ctx["user_acl_count"] = sum(1 for r in indexed if (r.get("allowed_users_count", 0) or 0) > 0)
            wb_ctx["group_acl_count"] = sum(1 for r in indexed if (r.get("allowed_groups_count", 0) or 0) > 0)
            wb_ctx["silent_count"] = sum(1 for r in indexed
                                         if (r.get("allowed_users_count", 0) or 0) == 0
                                         and (r.get("allowed_groups_count", 0) or 0) == 0
                                         and not r.get("allow_anonymous")
                                         and not r.get("allow_all_ds_users"))
            wb_ctx["visible_count"] = len(indexed) - wb_ctx["silent_count"]

            sheet_per_doc(wb, results)
            sheet_coverage_by_type(wb, results)
            sheet_acl_summary(wb, results, wb_ctx)
            sheet_silent_docs(wb, results)

    # ---------- 5. Per-user access checks ----------
    if args.user_email:
        items = load_batch_file(args.batch_file) if Path(args.batch_file).exists() else []
        for email in args.user_email:
            print(f"\n[5] Per-user access check: {email}")
            sheet_user_access(wb, email, items, session)

    # ---------- Status lines for Summary sheet ----------
    status_lines = []

    # Indexing health
    if wb_ctx.get("indexed_pct", 0) >= 99:
        status_lines.append({"label": "Indexing", "verdict": "HEALTHY",
                             "fg": GREEN_TXT, "bg": GREEN,
                             "detail": f"{wb_ctx['indexed_count']}/{wb_ctx['expected_total']} items confirmed in index"})
    elif wb_ctx.get("indexed_pct", 0) >= 80:
        status_lines.append({"label": "Indexing", "verdict": "PARTIAL",
                             "fg": YELLOW_TXT, "bg": YELLOW,
                             "detail": f"Only {wb_ctx.get('indexed_count', 0)}/{wb_ctx.get('expected_total', 0)} items in index. See Per-Doc sheet."})
    elif "indexed_count" in wb_ctx:
        status_lines.append({"label": "Indexing", "verdict": "FAILING",
                             "fg": RED_TXT, "bg": RED,
                             "detail": f"Only {wb_ctx['indexed_count']}/{wb_ctx['expected_total']} items. Re-run sync."})

    # ACL story
    if wb_ctx.get("silent_count", 0) == 0 and "indexed_count" in wb_ctx:
        status_lines.append({"label": "ACL Coverage", "verdict": "FULL",
                             "fg": GREEN_TXT, "bg": GREEN,
                             "detail": "All indexed docs have ACLs"})
    elif wb_ctx.get("silent_count", 0) > 0:
        pct = wb_ctx["silent_count"] / max(wb_ctx.get("indexed_count", 1), 1) * 100
        status_lines.append({"label": "ACL Coverage", "verdict": "INTENTIONAL GAPS" if pct < 70 else "SPARSE",
                             "fg": YELLOW_TXT, "bg": YELLOW,
                             "detail": f"{wb_ctx['silent_count']} docs ({pct:.0f}%) have no ACL — invisible in search by strict Vault parity"})

    # Run history
    if wb_ctx.get("bulk_history") or wb_ctx.get("proc_history"):
        status_lines.append({"label": "Run History", "verdict": "OK",
                             "fg": GREEN_TXT, "bg": GREEN,
                             "detail": "Upload and/or processing history present"})
    else:
        status_lines.append({"label": "Run History", "verdict": "EMPTY",
                             "fg": YELLOW_TXT, "bg": YELLOW,
                             "detail": "No history yet. If docs are in index but not searchable, check Admin Console to confirm datasource is enabled for search."})

    wb_ctx["status_lines"] = status_lines

    # ---------- Build Summary + Actions sheets ----------
    sheet_summary(wb, wb_ctx)
    sheet_actions(wb, wb_ctx)

    # Save
    wb.save(args.output)
    print(f"\n{'=' * 80}")
    print(f"REPORT SAVED: {args.output}")
    print(f"{'=' * 80}")


if __name__ == "__main__":
    main()